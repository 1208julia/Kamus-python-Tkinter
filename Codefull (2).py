#------------------------------------kode modifikasi lema dan pencarian biasa digabung fix---------------------------------------------------------#
import tkinter as tk
import openpyxl
import re

def load_data(file_path):
    data = {}
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            key = str(row[0]).strip()
            value = str(row[1]).strip()

            if key in data:
                data[key].append(value)
            else:
                data[key] = [value]
    except Exception as e:
        print(f"Error saat memuat data dari file: {e}")
        data = {}

    return data

def cari_kataDalamKalimat(kata, data):
    kataDalamKalimat = None
    for word, values in data.items():
        for value in values:
            if kata in value:
                kataDalamKalimat = word
                break  # Hanya temukan kecocokan pertama
    return kataDalamKalimat

def cari_lema(kata, data):
    lema = None
    for word, values in data.items():
        for value in values:
            if kata in value:
                lema = word
                break  # Hanya temukan kecocokan pertama
    return lema


def get_derivatives(keyword, data, use_kataDalamKalimat=False):
    derivatives = []
                
    if use_kataDalamKalimat:
        for word, values in data.items():
            for value in values:
                if keyword in value:
                    derivatives.append(word)
                    break   

    # Cari kata kunci terlebih dahulu
    if keyword in data:
        derivatives.append(keyword)

    # Cari kata lain yang sesuai dengan pencarian biasa
    for word in data.keys():
        if word.startswith(keyword.lower()) and word not in derivatives:
            derivatives.append(word)

    # Jika kata kunci tidak ditemukan, cari menggunakan pola regex
    if not derivatives:
        pattern = f"; {keyword} (v|n|a|adv|pron|p)"
        for word, values in data.items():
            for value in values:
                if re.search(pattern, value):
                    derivatives.append(word)
                    break

    return derivatives

def show_meaning(selected_word):
    meaning_text.config(state=tk.NORMAL)
    meaning_text.delete(1.0, tk.END)

    values = data.get(selected_word, ["Kata tidak ditemukan dalam kamus."])
    formatted_values = "\n\n".join(values)  # Gabungkan nilai-nilai menjadi satu string dengan pemisah berupa baris baru
    # Jika ada titik koma, ganti dengan baris baru untuk menampilkan nilai berikutnya di bawahnya
    #Label Kelas Kata
    formatted_values = formatted_values.replace(" n ", " [Nomina] ")
    formatted_values = formatted_values.replace(" a ", " [Adjektiva] ")
    formatted_values = formatted_values.replace(" v ", " [Verba] ")
    formatted_values = formatted_values.replace(" p ", " [Partikel] ")
    formatted_values = formatted_values.replace(" pron ", " [Pronomia] ")
    formatted_values = formatted_values.replace(" num ", " [Numeralia] ")
    formatted_values = formatted_values.replace(" adv ", " [Adverbia] ")
    formatted_values = formatted_values.replace(" ki ", " [Kiasan] ")
    
    formatted_values = formatted_values.replace(" ; ", "; ")
    formatted_values = formatted_values.replace("; --", "\n\t--")
    formatted_values = formatted_values.replace("; ~", "\n\t~")
    formatted_values = formatted_values.replace("; ", "\n\n")
    formatted_values = formatted_values.replace(".10", "\n\t10.")
    formatted_values = formatted_values.replace(".11", "\n\t11.")
    formatted_values = formatted_values.replace(".12", "\n\t12.")
    formatted_values = formatted_values.replace(".13", "\n\t13.")
    formatted_values = formatted_values.replace(".14", "\n\t14.")
    formatted_values = formatted_values.replace(".15", "\n\t15.")
    formatted_values = formatted_values.replace(".1", "\n\t1.")
    formatted_values = formatted_values.replace(".2", "\n\t2.")
    formatted_values = formatted_values.replace(".3", "\n\t3.")
    formatted_values = formatted_values.replace(".4", "\n\t4.")
    formatted_values = formatted_values.replace(".5", "\n\t5.")
    formatted_values = formatted_values.replace(".6", "\n\t6.")
    formatted_values = formatted_values.replace(".7", "\n\t7.")
    formatted_values = formatted_values.replace(".8", "\n\t8.")
    formatted_values = formatted_values.replace(".9", "\n\t9.")

    
    #Label Ragam Bahasa
    formatted_values = formatted_values.replace(" ark ", " [arkais] ")
    formatted_values = formatted_values.replace(" cak ", " [ragam cakapan] ")
    formatted_values = formatted_values.replace(" hor ", " [ragam hormat] ")
    formatted_values = formatted_values.replace(" kas ", " [kasar] ")
    formatted_values = formatted_values.replace(" kl ", " [klasik] ")
    
    #Bahasa Daerah
    formatted_values = formatted_values.replace(" Bl ", " [Bali] ")
    formatted_values = formatted_values.replace(" Bt ", " [Batak] ")
    formatted_values = formatted_values.replace(" Dy ", " [Sayak] ")
    formatted_values = formatted_values.replace(" Jw ", " [Jawa] ")
    formatted_values = formatted_values.replace(" Lp ", " [Lampung] ")
    formatted_values = formatted_values.replace(" Mdr ", " [Madura] ")
    formatted_values = formatted_values.replace(" Mk ", " [Minangkabau] ")
    formatted_values = formatted_values.replace(" Mn ", " [Minahasa] ")
    formatted_values = formatted_values.replace(" Mnd ", " [Menado] ")
    formatted_values = formatted_values.replace(" Plb ", " [Palembang] ")
    formatted_values = formatted_values.replace(" Sd ", " [Sunda] ")
    
    #Dialek  
    formatted_values = formatted_values.replace(" Jk ", " [Melayu Jakarta] ") 
    formatted_values = formatted_values.replace(" Mal ", " [Melayu Malaysia] ")

    #Bahasa Asing
    formatted_values = formatted_values.replace(" Ar ", " [Arab] ")
    formatted_values = formatted_values.replace(" Bld ", " [Belanda] ")
    formatted_values = formatted_values.replace(" Cn ", " [Cina] ")
    formatted_values = formatted_values.replace(" Ing ", " [Inggris] ")
    formatted_values = formatted_values.replace(" It ", " [Italia] ")
    formatted_values = formatted_values.replace(" Jm ", " [Jerman] ")
    formatted_values = formatted_values.replace(" Jp ", " [Jepang] ")
    formatted_values = formatted_values.replace(" Lt ", " [Latin] ")
    formatted_values = formatted_values.replace(" Par ", " [Parsi] ")
    formatted_values = formatted_values.replace(" Prt ", " [Portugis] ")
    formatted_values = formatted_values.replace(" Skot ", " [Skotlandia] ")
    formatted_values = formatted_values.replace(" Skt ", " [Sanskerta] ")
    formatted_values = formatted_values.replace(" Sp ", " [Spanyol] ")
    formatted_values = formatted_values.replace(" Yn ", " [Yunani] ")

    #Bahasa Kehidupan dan bidang Ilmu
    formatted_values = formatted_values.replace(" Adm ", " [Administrasi dan Kepegawaian] ")
    formatted_values = formatted_values.replace(" Anat ", " [Anatomi] ")
    formatted_values = formatted_values.replace(" Antr ", " [Antropologi] ")
    formatted_values = formatted_values.replace(" Ark ", " [Arkeologi] ")
    formatted_values = formatted_values.replace(" Ars ", " [Arsitektur] ")
    formatted_values = formatted_values.replace(" Astrol ", " [Astrologi] ")
    formatted_values = formatted_values.replace(" Astron ", " [Astronomi] ")
    formatted_values = formatted_values.replace(" Bakt ", " [Bakteriologi] ")
    formatted_values = formatted_values.replace(" Bio ", " [Biologi] ")
    formatted_values = formatted_values.replace(" Bot ", " [Botani] ")
    formatted_values = formatted_values.replace(" Bud ", " [Agama Budha] ")
    formatted_values = formatted_values.replace(" Dag ", " [Perdagangan] ")
    formatted_values = formatted_values.replace(" Dem ", " [Demografi] ")
    formatted_values = formatted_values.replace(" Dik ", " [Pendidikan] ")
    formatted_values = formatted_values.replace(" Dirg ", " [Kedirgantaraan] ")
    formatted_values = formatted_values.replace(" Dok ", " [Kedokteran dan Fisiologi] ")
    formatted_values = formatted_values.replace(" Ek ", " [Ekonomi dan Keuangan] ")
    formatted_values = formatted_values.replace(" El ", " [Elektronika (Kelistrikan dan Teknik Elektronika)] ")
    formatted_values = formatted_values.replace(" Ent ", " [Entomologi] ")
    formatted_values = formatted_values.replace(" Far ", " [Farmasi] ")
    formatted_values = formatted_values.replace(" Fil ", " [Filsafat] ")
    formatted_values = formatted_values.replace(" Filol ", " [Folologi] ")
    formatted_values = formatted_values.replace(" Fis ", " [Fisika] ")
    formatted_values = formatted_values.replace(" Geo ", " [Geografi dan Geologi] ")
    formatted_values = formatted_values.replace(" Graf ", " [Grafika] ")
    formatted_values = formatted_values.replace(" Hid ", " [Hidrologi] ")
    formatted_values = formatted_values.replace(" Hidm ", " [Hidrometeorologi] ")
    formatted_values = formatted_values.replace(" Hin ", " [Agama Hindu] ")
    formatted_values = formatted_values.replace(" Hub ", " [Perhubungan] ")
    formatted_values = formatted_values.replace(" Huk ", " [Hukum] ")
    formatted_values = formatted_values.replace(" Hut ", " [Kehutanan] ")
    formatted_values = formatted_values.replace(" Ikn ", " [Perikanan] ")
    formatted_values = formatted_values.replace(" Idt ", " [Perindustrian dan Kerajinan] ")
    formatted_values = formatted_values.replace(" Isl ", " [Agama Islam] ")
    formatted_values = formatted_values.replace(" Kap ", " [Perkapalan] ")
    formatted_values = formatted_values.replace(" Kat ", " [Agama Katolik] ")
    formatted_values = formatted_values.replace(" Kim ", " [Kimia] ")
    formatted_values = formatted_values.replace(" Kom ", " [Ilmu Komunikasi (Publisistik dan Jurnalistik)] ")
    formatted_values = formatted_values.replace(" Komp ", " [Komputer] ")
    formatted_values = formatted_values.replace(" Kris ", " [Agama Kristen] ")
    formatted_values = formatted_values.replace(" Lay ", " [Pelayaran] ")
    formatted_values = formatted_values.replace(" Ling ", " [Linguistik] ")
    formatted_values = formatted_values.replace(" Man ", " [Manajemen] ")
    formatted_values = formatted_values.replace(" Mat ", " [Matematika] ")
    formatted_values = formatted_values.replace(" Mek ", " [Mekanika] ")
    formatted_values = formatted_values.replace(" Met ", " [Meteorologi] ")
    formatted_values = formatted_values.replace(" Metal ", " [Metalurgi] ")
    formatted_values = formatted_values.replace(" Mik ", " [Mikologi] ")
    formatted_values = formatted_values.replace(" Mil ", " [Kemiliteran] ")
    formatted_values = formatted_values.replace(" Min ", " [Mineralogi] ")
    formatted_values = formatted_values.replace(" Mus ", " [Musik] ")
    formatted_values = formatted_values.replace(" Olr ", " [Olahraga] ")
    formatted_values = formatted_values.replace(" Pet ", " [Petrologi serta Minyak dan Gas Bumi] ")
    formatted_values = formatted_values.replace(" Pol ", " [Politik dan Pemerintahan] ")
    formatted_values = formatted_values.replace(" Psl ", " [Psikologi] ")
    formatted_values = formatted_values.replace(" Psi ", " [Psikologi] ")
    formatted_values = formatted_values.replace(" Sas ", " [Susastra (Sastra)] ")
    formatted_values = formatted_values.replace(" Sen ", " [Kesenian] ")
    formatted_values = formatted_values.replace(" Sos ", " [Sosiologi] ")
    formatted_values = formatted_values.replace(" Stat ", " [Statistik] ")
    formatted_values = formatted_values.replace(" Tan ", " [Pertanian] ")
    formatted_values = formatted_values.replace(" Tas ", " [Tasawuf] ")
    formatted_values = formatted_values.replace(" Tek ", " [Teknik] ")
    formatted_values = formatted_values.replace(" Telekom ", " [Telekomunikasi] ")
    formatted_values = formatted_values.replace(" Terb ", " [Penerbangan] ")
    formatted_values = formatted_values.replace(" Tern ", " [Peternakan] ")
    formatted_values = formatted_values.replace(" Zool ", " [Zoologi] ")

    # Update meaning_text properties
    meaning_text.config(state=tk.NORMAL, wrap=tk.NONE, height=20, width=70)
    meaning_text.delete(1.0, tk.END)
    meaning_text.insert(tk.END, formatted_values)
    meaning_text.config(state=tk.DISABLED)
    meaning_text.insert(tk.END, formatted_values)
    meaning_text.config(state=tk.DISABLED)
    

def on_listbox_select(event):
    selected_indices = listbox.curselection()
    if not selected_indices:
        return
    selected_word = listbox.get(selected_indices[0])
    # Cari kata yang dipilih dalam data dan tampilkan maknanya menggunakan show_meaning
    if selected_word in data:
        show_meaning(selected_word)
    else:
        # Jika kata tidak ada dalam data, tampilkan pesan kesalahan
        show_meaning("Kata tidak ditemukan dalam kamus.")

def on_search():
    keyword = search_entry.get().strip()
    if keyword:
        use_kataDalamKalimat = use_kataDalamKalimat_var.get()
        derivatives = get_derivatives(keyword, data, use_kataDalamKalimat=use_kataDalamKalimat)
        listbox.delete(0, tk.END)
        for word in derivatives:
            listbox.insert(tk.END, word)
        
        # Cek apakah ada hasil pencarian (derivatives) atau tidak
        if derivatives:
            # Tampilkan arti kata dari kata kunci paling atas di listbox
            selected_word = derivatives[0]
            show_meaning(selected_word)
        else:
            # Jika kata tidak ditemukan dalam data, hapus teks di meaning_text
            show_meaning("Kata tidak ditemukan dalam kamus.")


def create_label_frame(parent, text):
    label_frame = tk.LabelFrame(parent, text=text, padx=10, pady=10)
    label_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
    return label_frame

if __name__ == "__main__":
    data = load_data("D:\PKL\PROJEK PKL KODING AKADEMI\Bulan 1\DataAkhirKBBI.xlsx")  # Ganti "path_file_excel.xlsx" dengan path file Excel Anda

    window = tk.Tk()
    window.title("PROJEK PKL BULAN 1")
    window.geometry("860x560")

    main_label_frame = create_label_frame(window, "")

    intro_label = tk.Label(window, text="PROGRAM KAMUS BESAR BAHASA INDONESIA", font=("Arial", 12, "bold"))
    intro_label.place(x=215, y=25, height=15)

    label_frame1 = create_label_frame(main_label_frame, "")
    label_frame1.place(x=15, y=25, width=790)

    search_label = tk.Label(label_frame1, text="Masukkan kata:")
    search_label.grid(row=0, column=0, columnspan=3, padx=1, pady=1)
    search_entry = tk.Entry(label_frame1, width=125)
    search_entry.grid(row=1, column=0, columnspan=3, padx=1, pady=1)

    use_kataDalamKalimat_var = tk.IntVar()
    use_kataDalamKalimat_checkbutton = tk.Checkbutton(label_frame1, text="kata inputan dalam contoh arti kata", variable=use_kataDalamKalimat_var)
    use_kataDalamKalimat_checkbutton.grid(row=2, column=1, padx=5, pady=5)
    search_button = tk.Button(label_frame1, text="Cari", command=on_search)
    search_button.grid(row=2, column=2, padx=5, pady=5)
 
    label_frame2 = create_label_frame(main_label_frame, "")
    label_frame2.place(x=15, y=150)

    listbox_scrollbar = tk.Scrollbar(label_frame2, orient=tk.VERTICAL)
    listbox = tk.Listbox(label_frame2, width=20, height=21, yscrollcommand=listbox_scrollbar.set)
    listbox.grid(row=0, column=0, sticky="nsew")
    listbox_scrollbar.config(command=listbox.yview)
    listbox_scrollbar.grid(row=0, column=1, sticky="ns")
    
    label_frame3 = create_label_frame(main_label_frame, "")
    label_frame3.place(x=200, y=150)
    # Tambahkan komponen-komponen ke dalam label frame 3
    text_scrollbar_y = tk.Scrollbar(label_frame3, orient=tk.VERTICAL)
    text_scrollbar_x = tk.Scrollbar(label_frame3, orient=tk.HORIZONTAL)
    meaning_text = tk.Text(label_frame3, wrap=tk.WORD, width=70, height=20, padx=1, pady=1, state=tk.DISABLED, yscrollcommand=text_scrollbar_y.set, xscrollcommand=text_scrollbar_x.set)
    meaning_text.grid(row=0, column=0, sticky="nsew")
    text_scrollbar_y.config(command=meaning_text.yview)
    text_scrollbar_x.config(command=meaning_text.xview)
    text_scrollbar_y.grid(row=0, column=1, sticky="ns")
    text_scrollbar_x.grid(row=1, column=0, sticky="ew")

    meaning_text.config(xscrollcommand=text_scrollbar_x.set)

    listbox.bind("<<ListboxSelect>>", on_listbox_select)

    main_label_frame.grid_columnconfigure(1, weight=1)
    main_label_frame.grid_columnconfigure(2, weight=1)

    window.mainloop()



# #------------------------------------kode modifikasi lema dan pencarian biasa digabung---------------------------------------------------------#
# import tkinter as tk
# import openpyxl
# import re

# def load_data(file_path):
#     data = {}
#     try:
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook.active
#         for row in sheet.iter_rows(values_only=True):
#             key = str(row[0]).strip()
#             value = str(row[1]).strip()

#             if key in data:
#                 data[key].append(value)
#             else:
#                 data[key] = [value]
#     except Exception as e:
#         print(f"Error saat memuat data dari file: {e}")
#         data = {}

#     return data

# def cari_kataDalamKalimat(kata, data):
#     kataDalamKalimat = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 kataDalamKalimat = word
#                 break  # Hanya temukan kecocokan pertama
#     return kataDalamKalimat

# def cari_lema(kata, data):
#     lema = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 lema = word
#                 break  # Hanya temukan kecocokan pertama
#     return lema


# def get_derivatives(keyword, data, use_kataDalamKalimat=False):
#     derivatives = []
                
#     if use_kataDalamKalimat:
#         for word, values in data.items():
#             for value in values:
#                 if keyword in value:
#                     derivatives.append(word)
#                     break   

#     # Cari kata kunci terlebih dahulu
#     if keyword in data:
#         derivatives.append(keyword)

#     # Cari kata lain yang sesuai dengan pencarian biasa
#     for word in data.keys():
#         if word.startswith(keyword.lower()) and word not in derivatives:
#             derivatives.append(word)

#     # Jika kata kunci tidak ditemukan, cari menggunakan pola regex
#     if not derivatives:
#         pattern = f"; {keyword} (v|n|a|adv|pron|p|#)"
#         for word, values in data.items():
#             for value in values:
#                 if re.search(pattern, value):
#                     derivatives.append(word)
#                     break

#     return derivatives


# def show_meaning(selected_word):
#     meaning_text.config(state=tk.NORMAL)
#     meaning_text.delete(1.0, tk.END)

#     values = data.get(selected_word, ["Kata tidak ditemukan dalam kamus."])
#     formatted_values = "\n\n".join(values)  # Gabungkan nilai-nilai menjadi satu string dengan pemisah berupa baris baru
#     # Jika ada titik koma, ganti dengan baris baru untuk menampilkan nilai berikutnya di bawahnya
#     formatted_values = formatted_values.replace("; --", "\n\t--")
#     formatted_values = formatted_values.replace("; ~", "\n\t~")
#     formatted_values = formatted_values.replace("# ", "\n ")
#     formatted_values = formatted_values.replace("; ", "\n\n")
    
#     #Label Kelas Kata
#     formatted_values = formatted_values.replace(" n ", " [Nomina] ")
#     formatted_values = formatted_values.replace(" a ", " [Adjektiva] ")
#     formatted_values = formatted_values.replace(" v ", " [Verba] ")
#     formatted_values = formatted_values.replace(" p ", " [Partikel] ")
#     formatted_values = formatted_values.replace(" pron ", " [Pronomia] ")
#     formatted_values = formatted_values.replace(" num ", " [Numerelia] ")
#     formatted_values = formatted_values.replace(" adv ", " [Adverbia] ")
    
#     #Label Ragam Bahasa
#     formatted_values = formatted_values.replace(" ark ", " [arkais] ")
#     formatted_values = formatted_values.replace(" cak ", " [ragam cakapan] ")
#     formatted_values = formatted_values.replace(" hor ", " [ragam hormat] ")
#     formatted_values = formatted_values.replace(" kas ", " [kasar] ")
#     formatted_values = formatted_values.replace(" kl ", " [klasik] ")
    
#     #Bahasa Daerah
#     formatted_values = formatted_values.replace(" Bl ", " [Bali] ")
#     formatted_values = formatted_values.replace(" Bt ", " [Batak] ")
#     formatted_values = formatted_values.replace(" Dy ", " [Sayak] ")
#     formatted_values = formatted_values.replace(" Jw ", " [Jawa] ")
#     formatted_values = formatted_values.replace(" Lp ", " [Lampung] ")
#     formatted_values = formatted_values.replace(" Mdr ", " [Madura] ")
#     formatted_values = formatted_values.replace(" Mk ", " [Minangkabau] ")
#     formatted_values = formatted_values.replace(" Mn ", " [Minahasa] ")
#     formatted_values = formatted_values.replace(" Mnd ", " [Menado] ")
#     formatted_values = formatted_values.replace(" Plb ", " [Palembang] ")
#     formatted_values = formatted_values.replace(" Sd ", " [Sunda] ")
    
#     #Dialek  
#     formatted_values = formatted_values.replace(" Jk ", " [Melayu Jakarta] ") 
#     formatted_values = formatted_values.replace(" Mal ", " [Melayu Malaysia] ")

#     #Bahasa Asing
#     formatted_values = formatted_values.replace(" Ar ", " [Arab] ")
#     formatted_values = formatted_values.replace(" Bld ", " [Belanda] ")
#     formatted_values = formatted_values.replace(" Cn ", " [Cina] ")
#     formatted_values = formatted_values.replace(" Ing ", " [Inggris] ")
#     formatted_values = formatted_values.replace(" It ", " [Italia] ")
#     formatted_values = formatted_values.replace(" Jm ", " [Jerman] ")
#     formatted_values = formatted_values.replace(" Jp ", " [Jepang] ")
#     formatted_values = formatted_values.replace(" Lt ", " [Latin] ")
#     formatted_values = formatted_values.replace(" Par ", " [Parsi] ")
#     formatted_values = formatted_values.replace(" Prt ", " [Portugis] ")
#     formatted_values = formatted_values.replace(" Skot ", " [Skotlandia] ")
#     formatted_values = formatted_values.replace(" Skt ", " [Sanskerta] ")
#     formatted_values = formatted_values.replace(" Sp ", " [Spanyol] ")
#     formatted_values = formatted_values.replace(" Yn ", " [Yunani] ")

#     #Bahasa Kehidupan dan bidang Ilmu
#     formatted_values = formatted_values.replace(" Adm ", "[Administrasi dan Kepegawaian]")
#     formatted_values = formatted_values.replace(" Anat ", "[Anatomi]")
#     formatted_values = formatted_values.replace(" Antr ", "[Antropologi]")
#     formatted_values = formatted_values.replace(" Ark ", "[Arkeologi]")
#     formatted_values = formatted_values.replace(" Ars ", "[Arsitektur]")
#     formatted_values = formatted_values.replace(" Astrol ", "[Astrologi]")
#     formatted_values = formatted_values.replace(" Astron ", "[Astronomi]")
#     formatted_values = formatted_values.replace(" Bakt ", "[Bakteriologi]")
#     formatted_values = formatted_values.replace(" Bio ", "[Biologi]")
#     formatted_values = formatted_values.replace(" Bot ", "[Botani]")
#     formatted_values = formatted_values.replace(" Bud ", "[Agama Budha]")
#     formatted_values = formatted_values.replace(" Dag ", "[Perdagangan]")
#     formatted_values = formatted_values.replace(" Dem ", "[Demografi]")
#     formatted_values = formatted_values.replace(" Dik ", "[Pendidikan]")
#     formatted_values = formatted_values.replace(" Dirg ", "[Kedirgantaraan]")
#     formatted_values = formatted_values.replace(" Dok ", "[Kedokteran dan Fisiologi]")
#     formatted_values = formatted_values.replace(" Ek ", "[Ekonomi dan Keuangan]")
#     formatted_values = formatted_values.replace(" El ", "[Elektronika (Kelistrikan dan Teknik Elektronika)]")
#     formatted_values = formatted_values.replace(" Ent ", "[Entomologi]")
#     formatted_values = formatted_values.replace(" Far ", "[Farmasi]")
#     formatted_values = formatted_values.replace(" Fil ", "[Filsafat]")
#     formatted_values = formatted_values.replace(" Filol ", "[Folologi]")
#     formatted_values = formatted_values.replace(" Fis ", "[Fisika]")
#     formatted_values = formatted_values.replace(" Geo ", "[Geografi dan Geologi]")
#     formatted_values = formatted_values.replace(" Graf ", "[Grafika]")
#     formatted_values = formatted_values.replace(" Hid ", "[Hidrologi]")
#     formatted_values = formatted_values.replace(" Hidm ", "[Hidrometeorologi]")
#     formatted_values = formatted_values.replace(" Hin ", "[Agama Hindu]")
#     formatted_values = formatted_values.replace(" Hub ", "[Perhubungan]")
#     formatted_values = formatted_values.replace(" Huk ", "[Hukum]")
#     formatted_values = formatted_values.replace(" Hut ", "[Kehutanan]")
#     formatted_values = formatted_values.replace(" Ikn ", "[Perikanan]")
#     formatted_values = formatted_values.replace(" Idt ", "[Perindustrian dan Kerajinan]")
#     formatted_values = formatted_values.replace(" Isl ", "[Agama Islam]")
#     formatted_values = formatted_values.replace(" Kap ", "[Perkapalan]")
#     formatted_values = formatted_values.replace(" Kat ", "[Agama Katolik]")
#     formatted_values = formatted_values.replace(" Kim ", "[Kimia]")
#     formatted_values = formatted_values.replace(" Kom ", "[Ilmu Komunikasi (Publisistik dan Jurnalistik)]")
#     formatted_values = formatted_values.replace(" Komp ", "[Komputer]")
#     formatted_values = formatted_values.replace(" Kris ", "[Agama Kristen]")
#     formatted_values = formatted_values.replace(" Lay ", "[Pelayaran]")
#     formatted_values = formatted_values.replace(" Ling ", "[Linguistik]")
#     formatted_values = formatted_values.replace(" Man ", "[Manajemen]")
#     formatted_values = formatted_values.replace(" Mat ", "[Matematika]")
#     formatted_values = formatted_values.replace(" Mek ", "[Mekanika]")
#     formatted_values = formatted_values.replace(" Met ", "[Meteorologi]")
#     formatted_values = formatted_values.replace(" Metal ", "[Metalurgi]")
#     formatted_values = formatted_values.replace(" Mik ", "[Mikologi]")
#     formatted_values = formatted_values.replace(" Mil ", "[Kemiliteran]")
#     formatted_values = formatted_values.replace(" Min ", "[Mineralogi]")
#     formatted_values = formatted_values.replace(" Mus ", "[Musik]")
#     formatted_values = formatted_values.replace(" Olr ", "[Olahraga]")
#     formatted_values = formatted_values.replace(" Pet ", "[Petrologi serta Minyak dan Gas Bumi]")
#     formatted_values = formatted_values.replace(" Pol ", "[Politik dan Pemerintahan]")
#     formatted_values = formatted_values.replace(" Psl ", "[Psikologi]")
#     formatted_values = formatted_values.replace(" Sas ", "[Susastra (Sastra)]")
#     formatted_values = formatted_values.replace(" Sen ", "[Kesenian]")
#     formatted_values = formatted_values.replace(" Sos ", "[Sosiologi]")
#     formatted_values = formatted_values.replace(" Stat ", "[Statistik]")
#     formatted_values = formatted_values.replace(" Tan ", "[Pertanian]")
#     formatted_values = formatted_values.replace(" Tas ", "[Tasawuf]")
#     formatted_values = formatted_values.replace(" Tek ", "[Teknik]")
#     formatted_values = formatted_values.replace(" Telekom ", "[Telekomunikasi]")
#     formatted_values = formatted_values.replace(" Terb ", "[Penerbangan]")
#     formatted_values = formatted_values.replace(" Tern ", "[Peternakan]")
#     formatted_values = formatted_values.replace(" Zool ", "[Zoologi]")

    
#     meaning_text.insert(tk.END, formatted_values)
#     meaning_text.config(state=tk.DISABLED)


# def on_listbox_select(event):
#     selected_indices = listbox.curselection()
#     if not selected_indices:
#         return
#     selected_word = listbox.get(selected_indices[0])
#     # Cari kata yang dipilih dalam data dan tampilkan maknanya menggunakan show_meaning
#     if selected_word in data:
#         show_meaning(selected_word)
#     else:
#         # Jika kata tidak ada dalam data, tampilkan pesan kesalahan
#         show_meaning("Kata tidak ditemukan dalam kamus.")


# def on_search():
#     keyword = search_entry.get().strip()
#     if keyword:
#         use_kataDalamKalimat = use_kataDalamKalimat_var.get()
#         derivatives = get_derivatives(keyword, data, use_kataDalamKalimat=use_kataDalamKalimat)
#         listbox.delete(0, tk.END)
#         for word in derivatives:
#             listbox.insert(tk.END, word)


# def create_label_frame(parent, text):
#     label_frame = tk.LabelFrame(parent, text=text, padx=10, pady=10)
#     label_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
#     return label_frame

# if __name__ == "__main__":
#     data = load_data("D:\PKL\PROJEK PKL KODING AKADEMI\Bulan 1\coba9\data_plain5_updated.xlsx")  # Ganti "path_file_excel.xlsx" dengan path file Excel Anda
    
# #--------------------------------------TK INTER GRID-------------------------#
#     # Membuat jendela utama
#     window = tk.Tk()
#     window.title("PROJEK PKL BULAN 1")
#     window.geometry("860x560")

#     # Membuat label frame besar
#     main_label_frame = create_label_frame(window, "")

#     intro_label = tk.Label(window, text="PROGRAM KAMUS BESAR BAHASA INDONESIA", font=("Arial", 12, "bold"))
#     intro_label.place(x=215, y=25, height=15)

#     # Membuat label frame ke-1 di dalam label frame besar
#     label_frame1 = create_label_frame(main_label_frame, "")
#     label_frame1.place(x=15, y=25, width=790)
#     # Tambahkan komponen-komponen ke dalam label frame 1
#     search_label = tk.Label(label_frame1, text="Masukkan kata:")
#     search_label.grid(row=0, column=0, columnspan=3, padx=1, pady=1)
#     search_entry = tk.Entry(label_frame1, width=125)
#     search_entry.grid(row=1, column=0, columnspan=3, padx=1, pady=1)

#     # Tambahkan use_kataDalamKalimat_var sebagai variabel status untuk fitur "Kata dalam Kalimat" atau "Lema"
#     use_kataDalamKalimat_var = tk.IntVar()
#     use_kataDalamKalimat_checkbutton = tk.Checkbutton(label_frame1, text="Kata dalam Kalimat", variable=use_kataDalamKalimat_var)
#     use_kataDalamKalimat_checkbutton.grid(row=2, column=1, padx=5, pady=5)
#     search_button = tk.Button(label_frame1, text="Cari", command=on_search)
#     search_button.grid(row=2, column=2, padx=5, pady=5)
 
#     # Membuat label frame ke-2 di dalam label frame besar
#     label_frame2 = create_label_frame(main_label_frame, "")
#     label_frame2.place(x=15, y=150)
#     # Tambahkan komponen-komponen ke dalam label frame 2
#     listbox_scrollbar = tk.Scrollbar(label_frame2, orient=tk.VERTICAL)
#     listbox = tk.Listbox(label_frame2, width=20, height=20, yscrollcommand=listbox_scrollbar.set)
#     listbox.grid(row=0, column=0, sticky="nsew")
#     listbox_scrollbar.config(command=listbox.yview)
#     listbox_scrollbar.grid(row=0, column=1, sticky="ns")

#     # Membuat label frame ke-3 di dalam label frame besar
#     label_frame3 = create_label_frame(main_label_frame, "")
#     label_frame3.place(x=200, y=150)
#     # Tambahkan komponen-komponen ke dalam label frame 3
#     text_scrollbar_y = tk.Scrollbar(label_frame3, orient=tk.VERTICAL)
#     text_scrollbar_x = tk.Scrollbar(label_frame3, orient=tk.HORIZONTAL)
#     meaning_text = tk.Text(label_frame3, wrap=tk.WORD, width=70, height=20, padx=1, pady=1, state=tk.DISABLED, yscrollcommand=text_scrollbar_y.set, xscrollcommand=text_scrollbar_x.set)
#     meaning_text.grid(row=0, column=0, sticky="nsew")
#     text_scrollbar_y.config(command=meaning_text.yview)
#     text_scrollbar_x.config(command=meaning_text.xview)
#     text_scrollbar_y.grid(row=0, column=1, sticky="ns")
#     text_scrollbar_x.grid(row=1, column=0, sticky="ew")
    
#     meaning_text.config(xscrollcommand=text_scrollbar_x.set)

#     listbox.bind("<<ListboxSelect>>", on_listbox_select)

#     # Mengatur agar Label Frame 2 sejajar dengan Label Frame 3
#     main_label_frame.grid_columnconfigure(1, weight=1)
#     main_label_frame.grid_columnconfigure(2, weight=1)

#     window.mainloop()





# #------------------------------------kode awal---------------------------------------------------------#
# import tkinter as tk
# import openpyxl
# import re

# def load_data(file_path):
#     data = {}
#     try:
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook.active
#         for row in sheet.iter_rows(values_only=True):
#             key = str(row[0]).strip()
#             value = str(row[1]).strip()

#             if key in data:
#                 data[key].append(value)
#             else:
#                 data[key] = [value]
#     except Exception as e:
#         print(f"Error saat memuat data dari file: {e}")
#         data = {}

#     return data
 
# def cari_kataDalamKalimat(kata, data):
#     kataDalamKalimat = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 kataDalamKalimat = word
#                 break  # Hanya temukan kecocokan pertama
#     return kataDalamKalimat

# def cari_lema(kata, data):
#     lema = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 lema = word
#                 break  # Hanya temukan kecocokan pertama
#     return lema

# def get_derivatives(keyword, data, use_lema=False, use_kataDalamKalimat=False):
#     derivatives = []
    
#     if use_lema:
#         # Definisikan pola regex untuk mencocokkan pola yang diinginkan
#         pattern = f"; {keyword} (v|n|a|adv|pron|p|#)"

#         for word, values in data.items():
#             for value in values:
#                 if re.search(pattern, value):
#                     derivatives.append(word)
#                     break 
                
#     elif use_kataDalamKalimat:
#         for word, values in data.items():
#             for value in values:
#                 if keyword in value:
#                     derivatives.append(word)
#                     break   
#     else:
#         derivatives = [word for word in data.keys() if word.startswith(keyword.lower())]

#     return derivatives


# def show_meaning(selected_word):
#     meaning_text.config(state=tk.NORMAL)
#     meaning_text.delete(1.0, tk.END)

#     values = data.get(selected_word, ["Kata tidak ditemukan dalam kamus."])
#     formatted_values = "\n\n".join(values)  # Gabungkan nilai-nilai menjadi satu string dengan pemisah berupa baris baru

#     # Jika ada titik koma, ganti dengan baris baru untuk menampilkan nilai berikutnya di bawahnya
#     formatted_values = formatted_values.replace("; --", "\n\t--")
#     formatted_values = formatted_values.replace("; ~", "\n\t~")
#     formatted_values = formatted_values.replace("# ", "\n ")
#     formatted_values = formatted_values.replace("; ", "\n\n")

#     meaning_text.insert(tk.END, formatted_values)
#     meaning_text.config(state=tk.DISABLED)


# def on_listbox_select(event):
#     selected_word = listbox.get(listbox.curselection())
#     # Cari kata yang dipilih dalam data dan tampilkan maknanya menggunakan show_meaning
#     if selected_word in data:
#         show_meaning(selected_word)
#     else:
#         # Jika kata tidak ada dalam data, tampilkan pesan kesalahan
#         show_meaning("Kata tidak ditemukan dalam kamus.")

# def on_search():
#     keyword = search_entry.get().strip()
#     if keyword:
#         use_lema = use_lema_var.get()
#         use_kataDalamKalimat = use_kataDalamKalimat_var.get()
#         derivatives = get_derivatives(keyword, data, use_lema, use_kataDalamKalimat)
#         listbox.delete(0, tk.END)
#         for word in derivatives:
#             listbox.insert(tk.END, word)

# def create_label_frame(parent, text):
#     label_frame = tk.LabelFrame(parent, text=text, padx=10, pady=10)
#     label_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
#     return label_frame

# if __name__ == "__main__":
#     data = load_data("D:\PKL\PROJEK PKL KODING AKADEMI\Bulan 1\coba9\data_plain5_updated.xlsx")  # Ganti "path_file_excel.xlsx" dengan path file Excel Anda
    
# #--------------------------------------TK INTER GRID-------------------------#
#     # Membuat jendela utama
#     window = tk.Tk()
#     window.title("PROJEK PKL BULAN 1")
#     window.geometry("860x560")

#     # Membuat label frame besar
#     main_label_frame = create_label_frame(window, "")

#     intro_label = tk.Label(window, text="PROGRAM KAMUS BESAR BAHASA INDONESIA", font=("Arial", 12, "bold"))
#     intro_label.place(x=215, y=25, height=15)

#     # Membuat label frame ke-1 di dalam label frame besar
#     label_frame1 = create_label_frame(main_label_frame, "")
#     label_frame1.place(x=15, y=25, width=790)
#     # Tambahkan komponen-komponen ke dalam label frame 1
#     search_label = tk.Label(label_frame1, text="Masukkan kata:")
#     search_label.grid(row=0, column=0, columnspan=3, padx=1, pady=1)
#     search_entry = tk.Entry(label_frame1, width=125)
#     search_entry.grid(row=1, column=0, columnspan=3, padx=1, pady=1)
#     use_lema_var = tk.IntVar()
#     use_lema_checkbutton = tk.Checkbutton(label_frame1, text="Lema", variable=use_lema_var)
#     use_lema_checkbutton.grid(row=2, column=0, padx=5, pady=5)
#     # Tambahkan use_kataDalamKalimat_var sebagai variabel status untuk fitur "Kata dalam Kalimat" atau "Lema"
#     use_kataDalamKalimat_var = tk.IntVar()
#     use_kataDalamKalimat_checkbutton = tk.Checkbutton(label_frame1, text="Kata dalam Kalimat", variable=use_kataDalamKalimat_var)
#     use_kataDalamKalimat_checkbutton.grid(row=2, column=1, padx=5, pady=5)
#     search_button = tk.Button(label_frame1, text="Cari", command=on_search)
#     search_button.grid(row=2, column=2, padx=5, pady=5)
 
#     # Membuat label frame ke-2 di dalam label frame besar
#     label_frame2 = create_label_frame(main_label_frame, "")
#     label_frame2.place(x=15, y=150)
#     # Tambahkan komponen-komponen ke dalam label frame 2
#     listbox_scrollbar = tk.Scrollbar(label_frame2, orient=tk.VERTICAL)
#     listbox = tk.Listbox(label_frame2, width=20, height=20, yscrollcommand=listbox_scrollbar.set)
#     listbox.grid(row=0, column=0, sticky="nsew")
#     listbox_scrollbar.config(command=listbox.yview)
#     listbox_scrollbar.grid(row=0, column=1, sticky="ns")

#     # Membuat label frame ke-3 di dalam label frame besar
#     label_frame3 = create_label_frame(main_label_frame, "")
#     label_frame3.place(x=200, y=150)
#     # Tambahkan komponen-komponen ke dalam label frame 3
#     text_scrollbar = tk.Scrollbar(label_frame3, orient=tk.VERTICAL)
#     meaning_text = tk.Text(label_frame3, wrap=tk.WORD, width=70, height=20, padx=1, pady=1, state=tk.DISABLED, yscrollcommand=text_scrollbar.set)
#     meaning_text.grid(row=0, column=0, sticky="nsew")
#     text_scrollbar.config(command=meaning_text.yview)
#     text_scrollbar.grid(row=0, column=1, sticky="ns")

#     listbox.bind("<<ListboxSelect>>", on_listbox_select)

#     # Mengatur agar Label Frame 2 sejajar dengan Label Frame 3
#     main_label_frame.grid_columnconfigure(1, weight=1)
#     main_label_frame.grid_columnconfigure(2, weight=1)

#     window.mainloop()
    
    # def load_data(file_path):
#     data = {}
#     try:
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook.active
#         for row in sheet.iter_rows(values_only=True):
#             key = str(row[0]).strip()
#             value = str(row[1]).strip()

#             if key in data:
#                 data[key].append(value)
#             else:
#                 data[key] = [value]
#     except Exception as e:
#         print(f"Error loading data from file: {e}")
#         data = {}

#     return data
 
# def cari_lema(kata, data):
#     lema = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 lema = word
#                 break  # Only find the first match
#     return lema

# def get_derivatives(keyword, data, use_lema=False):
#     derivatives = []

#     if use_lema:
#         for word, values in data.items():
#             for value in values:
#                 if keyword in value:
#                     derivatives.append(word)
#                     break  # Only add the word once if the lemma appears multiple times in the values
#     else:
#         derivatives = [word for word in data.keys() if word.startswith(keyword.lower())]

#     return derivatives

# def show_meaning(selected_word):
#     meaning_text.config(state=tk.NORMAL)
#     meaning_text.delete(1.0, tk.END)

#     values = data.get(selected_word, ["Kata tidak ditemukan dalam kamus."])
#     formatted_values = "\n\n".join(values)  # Combine the values into one string with line breaks as separators

#     # If there are semicolons, replace them with line breaks to display the next value below it
#     formatted_values = formatted_values.replace("; --", "\n\t--")
#     formatted_values = formatted_values.replace("; ~", "\n\t~")
#     formatted_values = formatted_values.replace("; ", "\n ")
#     formatted_values = formatted_values.replace(";", "\n\n")

#     meaning_text.insert(tk.END, formatted_values)
#     meaning_text.config(state=tk.DISABLED)


# def on_listbox_select(event):
#     selected_word = listbox.get(listbox.curselection())
#     # Cari kata yang dipilih dalam data dan tampilkan maknanya menggunakan show_meaning
#     if selected_word in data:
#         show_meaning(selected_word)
#     else:
#         # Jika kata tidak ada dalam data, tampilkan pesan kesalahan
#         show_meaning("Kata tidak ditemukan dalam kamus.")

# def on_search():
#     keyword = search_entry.get().strip()
#     if keyword:
#         use_lema = use_lema_var.get()
#         derivatives = get_derivatives(keyword, data, use_lema)
#         listbox.delete(0, tk.END)
#         for word in derivatives:
#             listbox.insert(tk.END, word)
#---------------------------------------------------------




#------------------------------------kode awal---------------------------------------------------------#
# import tkinter as tk
# import openpyxl
# import re

# def load_data(file_path):
#     data = {}
#     try:
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook.active
#         for row in sheet.iter_rows(values_only=True):
#             key = str(row[0]).strip()
#             value = str(row[1]).strip()

#             if key in data:
#                 data[key].append(value)
#             else:
#                 data[key] = [value]
#     except Exception as e:
#         print(f"Error saat memuat data dari file: {e}")
#         data = {}

#     return data
 

# def cari_lema(kata, data):
#     lema = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 lema = word
#                 break  # Hanya temukan kecocokan pertama
#     return lema

# def get_derivatives(keyword, data, use_lema=False):
#     derivatives = []
    
#     if use_lema:
#         # Definisikan pola regex untuk mencocokkan pola yang diinginkan
#         pattern = f"; {keyword} (v|n|a|adv|pron|p|#)"

#         for word, values in data.items():
#             for value in values:
#                 if re.search(pattern, value):
#                     derivatives.append(word)
#                     break 
#     else:
#         derivatives = [word for word in data.keys() if word.startswith(keyword.lower())]

#     return derivatives

# def show_meaning(selected_word):
#     meaning_text.config(state=tk.NORMAL)
#     meaning_text.delete(1.0, tk.END)

#     values = data.get(selected_word, ["Kata tidak ditemukan dalam kamus."])
#     formatted_values = "\n\n".join(values)  # Gabungkan nilai-nilai menjadi satu string dengan pemisah berupa baris baru

#     # Jika ada titik koma, ganti dengan baris baru untuk menampilkan nilai berikutnya di bawahnya
#     formatted_values = formatted_values.replace("; --", "\n\t--")
#     formatted_values = formatted_values.replace("; ~", "\n\t~")
#     formatted_values = formatted_values.replace("# ", "\n ")
#     formatted_values = formatted_values.replace("; ", "\n\n")

#     meaning_text.insert(tk.END, formatted_values)
#     meaning_text.config(state=tk.DISABLED)


# def on_listbox_select(event):
#     selected_word = listbox.get(listbox.curselection())
#     # Cari kata yang dipilih dalam data dan tampilkan maknanya menggunakan show_meaning
#     if selected_word in data:
#         show_meaning(selected_word)
#     else:
#         # Jika kata tidak ada dalam data, tampilkan pesan kesalahan
#         show_meaning("Kata tidak ditemukan dalam kamus.")

# def on_search():
#     keyword = search_entry.get().strip()
#     if keyword:
#         use_lema = use_lema_var.get()
#         derivatives = get_derivatives(keyword, data, use_lema)
#         listbox.delete(0, tk.END)
#         for word in derivatives:
#             listbox.insert(tk.END, word)

# def create_label_frame(parent, text):
#     label_frame = tk.LabelFrame(parent, text=text, padx=10, pady=10)
#     label_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
#     return label_frame

# if __name__ == "__main__":
#     data = load_data("D:\PKL\PROJEK PKL KODING AKADEMI\Bulan 1\coba9\data_plain5_updated.xlsx")  # Ganti "path_file_excel.xlsx" dengan path file Excel Anda
    
# #--------------------------------------TK INTER GRID-------------------------#
#     # Membuat jendela utama
#     window = tk.Tk()
#     window.title("PROJEK PKL BULAN 1")
#     window.geometry("860x560")

#     # Membuat label frame besar
#     main_label_frame = create_label_frame(window, "")

#     intro_label = tk.Label(window, text="PROGRAM KAMUS BESAR BAHASA INDONESIA", font=("Arial", 12, "bold"))
#     intro_label.place(x=215, y=25)

#     # Membuat label frame ke-1 di dalam label frame besar
#     label_frame1 = create_label_frame(main_label_frame, "")
#     label_frame1.place(x=15, y=15)
#     # Tambahkan komponen-komponen ke dalam label frame 1
#     search_label = tk.Label(label_frame1, text="Masukkan kata:")
#     search_label.grid(row=0, column=0, columnspan=2, padx=1, pady=1)
#     search_entry = tk.Entry(label_frame1, width=125)
#     search_entry.grid(row=1, column=0, columnspan=2, padx=5, pady=5)
#     use_lema_var = tk.IntVar()
#     use_lema_checkbutton = tk.Checkbutton(label_frame1, text="Lema", variable=use_lema_var)
#     use_lema_checkbutton.grid(row=2, column=0, padx=5, pady=5)
#     search_button = tk.Button(label_frame1, text="Cari", command=on_search)
#     search_button.grid(row=2, column=1, padx=5, pady=5)

#     # Membuat label frame ke-2 di dalam label frame besar
#     label_frame2 = create_label_frame(main_label_frame, "")
#     label_frame2.place(x=15, y=150)
#     # Tambahkan komponen-komponen ke dalam label frame 2
#     listbox_scrollbar = tk.Scrollbar(label_frame2, orient=tk.VERTICAL)
#     listbox = tk.Listbox(label_frame2, width=20, height=20, yscrollcommand=listbox_scrollbar.set)
#     listbox.grid(row=0, column=0, sticky="nsew")
#     listbox_scrollbar.config(command=listbox.yview)
#     listbox_scrollbar.grid(row=0, column=1, sticky="ns")

#     # Membuat label frame ke-3 di dalam label frame besar
#     label_frame3 = create_label_frame(main_label_frame, "")
#     label_frame3.place(x=200, y=150)
#     # Tambahkan komponen-komponen ke dalam label frame 3
#     text_scrollbar = tk.Scrollbar(label_frame3, orient=tk.VERTICAL)
#     meaning_text = tk.Text(label_frame3, wrap=tk.WORD, width=70, height=20, padx=1, pady=1, state=tk.DISABLED, yscrollcommand=text_scrollbar.set)
#     meaning_text.grid(row=0, column=0, sticky="nsew")
#     text_scrollbar.config(command=meaning_text.yview)
#     text_scrollbar.grid(row=0, column=1, sticky="ns")

#     listbox.bind("<<ListboxSelect>>", on_listbox_select)

#     # Mengatur agar Label Frame 2 sejajar dengan Label Frame 3
#     main_label_frame.grid_columnconfigure(1, weight=1)
#     main_label_frame.grid_columnconfigure(2, weight=1)

#     window.mainloop()

#----------------------------------
# def cari_lema(kata, data):
#     if kata in data:
#         # Jika kata langsung cocok dengan kata dalam data, kembalikan kata tersebut
#         return kata
    
#     lema = None
#     for word, values in data.items():
#         for value in values:
#             if kata in value:
#                 lema = word
#                 break  # Hanya temukan kecocokan pertama
#     return lema

# def get_derivatives(keyword, data, use_lema=False, use_kataDalamKalimat=False):
#     derivatives = []
    
#     if use_lema:
#         # Cek terlebih dahulu apakah kata kunci ada dalam data
#         lema_result = cari_lema(keyword, data)
#         if lema_result:
#             derivatives.append(lema_result)
#         else:
#             # Jika tidak ada, cari menggunakan pola regex
#             pattern = f"; {keyword} (v|n|a|adv|pron|p|#)"

#             for word, values in data.items():
#                 for value in values:
#                     if re.search(pattern, value):
#                         derivatives.append(word)
#                         break 
#     elif use_kataDalamKalimat:
#         for word, values in data.items():
#             for value in values:
#                 if keyword in value:
#                     derivatives.append(word)
#                     break   
#     else:
#         derivatives = [word for word in data.keys() if word.startswith(keyword.lower())]

#     return derivatives
